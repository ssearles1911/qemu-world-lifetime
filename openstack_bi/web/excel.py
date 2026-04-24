"""Generic `.xlsx` exporter for any `ReportResult`.

Structure:
    Sheet "Data":
        Top: metadata block (one key: value per row).
        Blank row.
        Header row (bold, gray fill, frozen).
        Data rows.
        Auto-filter on every column.

    One sheet per chart in `result.charts`, rendered as a matplotlib PNG
    embedded at A1 alongside the raw series as a small table. Matplotlib
    is imported lazily so reports without charts don't pull it in.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, List

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from openstack_bi.reports.base import ChartSpec, ReportResult


def build(result: ReportResult) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for key, value in result.metadata.items():
        ws.append([f"{key}: {value}"])
    ws.append([f"generated_at: {datetime.utcnow().isoformat(timespec='seconds')}Z"])
    ws.append([])

    visible = [(k, label) for k, label in result.columns if not label.startswith("_")]
    header_row = ws.max_row + 1
    ws.append([label for _, label in visible])
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for col_idx in range(1, len(visible) + 1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="left")

    for r in result.rows:
        ws.append([r.get(k) for k, _ in visible])

    last_row = ws.max_row
    last_col_letter = get_column_letter(max(1, len(visible)))

    if last_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for col_idx in range(1, len(visible) + 1):
        col_letter = get_column_letter(col_idx)
        values = [ws.cell(row=r, column=col_idx).value for r in range(header_row, last_row + 1)]
        max_len = max((len(str(v)) for v in values if v is not None), default=10)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    for idx, chart in enumerate(result.charts, start=1):
        _add_chart_sheet(wb, chart, idx)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _add_chart_sheet(wb: Workbook, chart: ChartSpec, idx: int) -> None:
    safe_title = (chart.title or f"chart{idx}").replace("/", "-").replace(":", "-")[:31]
    sheet_name = safe_title or f"chart{idx}"
    ws = wb.create_sheet(title=sheet_name)

    png_bytes = _render_chart_png(chart)
    if png_bytes is not None:
        ws.add_image(XLImage(io.BytesIO(png_bytes)), "A1")

    start_row = 24  # leave space below the embedded chart image
    ws.cell(row=start_row, column=1, value=chart.x_label or "x")
    for s_idx, series in enumerate(chart.series, start=2):
        ws.cell(row=start_row, column=s_idx, value=series.get("label"))
    for i, cat in enumerate(chart.x_categories, start=1):
        ws.cell(row=start_row + i, column=1, value=cat)
        for s_idx, series in enumerate(chart.series, start=2):
            data: List[Any] = series.get("data") or []
            ws.cell(
                row=start_row + i, column=s_idx,
                value=data[i - 1] if i - 1 < len(data) else None,
            )


def _render_chart_png(chart: ChartSpec) -> "bytes | None":
    """Render a ChartSpec as a PNG via matplotlib. Returns None if
    matplotlib isn't installed — the chart tab still gets the raw data.
    """
    try:
        import matplotlib  # noqa: WPS433
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt  # noqa: WPS433
    except ImportError:
        return None

    fig, ax = plt.subplots(figsize=(9, 4.5))
    x = list(range(len(chart.x_categories)))

    if chart.kind == "line":
        for s in chart.series:
            ax.plot(x, s.get("data") or [], marker="o", label=s.get("label"))
    elif chart.kind == "stacked_bar":
        bottom = [0.0] * len(chart.x_categories)
        for s in chart.series:
            data = list(s.get("data") or [])
            # pad/trim to x length
            data = (data + [0.0] * len(chart.x_categories))[: len(chart.x_categories)]
            ax.bar(x, data, bottom=bottom, label=s.get("label"))
            bottom = [b + (d or 0) for b, d in zip(bottom, data)]
    else:  # default "bar" — grouped side-by-side
        n = max(1, len(chart.series))
        width = 0.8 / n
        for idx, s in enumerate(chart.series):
            offsets = [xi + (idx - (n - 1) / 2) * width for xi in x]
            ax.bar(offsets, s.get("data") or [], width=width, label=s.get("label"))

    ax.set_title(chart.title)
    ax.set_xlabel(chart.x_label)
    ax.set_ylabel(chart.y_label)
    ax.set_xticks(x)
    ax.set_xticklabels(chart.x_categories, rotation=30, ha="right", fontsize=8)
    if len(chart.series) > 1:
        ax.legend(fontsize=8)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110)
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()
