#!/usr/bin/env python3
"""Flask web UI for the QEMU lifetime report.

Routes:
    GET /                  — domain dropdown + days filter; renders the
                              grouped report when ?domain=... is set.
    GET /export.xlsx       — same params, returns an Excel workbook.

Run:
    python web.py
    QLR_HOST=0.0.0.0 QLR_PORT=8000 python web.py

DB connection comes from `core.py` env vars (OS_DB_*, KEYSTONE_DB,
NOVA_API_DB).
"""

import io
import os
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional

from flask import Flask, abort, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import core


app = Flask(__name__)


ALL_STATES_SENTINEL = "__all__"


def _parse_days(raw: Optional[str]) -> Optional[int]:
    if raw is None or raw == "":
        return None
    try:
        n = int(raw)
        return n if n >= 0 else None
    except ValueError:
        return None


def _parse_state(raw: Optional[str]) -> (Optional[List[str]], str):
    """Resolve the `state` query arg into (vm_states, selected_state).

    `vm_states` is what core.collect_report() expects: None means "any
    state", a list means "filter to these". `selected_state` is what
    the dropdown should show as picked (the sentinel or a state name).

    If no `state` param is sent (first visit / direct link), default to
    'active' — matches the project convention.
    """
    if raw is None:
        return list(core.DEFAULT_VM_STATES), core.DEFAULT_VM_STATES[0]
    if raw == ALL_STATES_SENTINEL or raw == "":
        return None, ALL_STATES_SENTINEL
    return [raw], raw


@app.route("/")
def index():
    domains = core.list_domains()
    domain_sel = request.args.get("domain", "").strip()
    days = _parse_days(request.args.get("days"))
    vm_states, selected_state = _parse_state(request.args.get("state"))

    context: Dict[str, Any] = {
        "domains": domains,
        "selected_domain": domain_sel,
        "days": "" if days is None else days,
        "lifecycle_actions": core.LIFECYCLE_ACTIONS,
        "common_vm_states": core.COMMON_VM_STATES,
        "selected_state": selected_state,
        "all_states_sentinel": ALL_STATES_SENTINEL,
        "domain": None,
        "projects": [],
        "rows_by_project": {},
        "total": 0,
        "error": None,
    }

    if domain_sel:
        report = core.collect_report(domain_sel, days, vm_states)
        if report["domain"] is None:
            context["error"] = f"Domain not found: {domain_sel}"
        else:
            by_project: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
            for r in report["rows"]:
                by_project[r["project_id"]].append(r)
            context["domain"] = report["domain"]
            context["projects"] = report["projects"]
            context["rows_by_project"] = by_project
            context["total"] = len(report["rows"])

    return render_template("index.html", **context)


@app.route("/export.xlsx")
def export_xlsx():
    domain_sel = request.args.get("domain", "").strip()
    if not domain_sel:
        abort(400, "Missing 'domain' parameter")
    days = _parse_days(request.args.get("days"))
    vm_states, selected_state = _parse_state(request.args.get("state"))

    report = core.collect_report(domain_sel, days, vm_states)
    if report["domain"] is None:
        abort(404, f"Domain not found: {domain_sel}")

    bio = _build_workbook(
        report["domain"], report["projects"], report["rows"], days, vm_states,
    )

    bits = [report["domain"]["name"], "qemu-lifetime"]
    if vm_states:
        bits.append("-".join(vm_states))
    else:
        bits.append("all-states")
    if days is not None:
        bits.append(f"{days}d")
    filename = "-".join(bits) + ".xlsx"

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _build_workbook(
    domain: Dict[str, Any],
    projects: List[Dict[str, Any]],
    rows: List[Dict[str, Any]],
    days: Optional[int],
    vm_states: Optional[List[str]],
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = (domain["name"] or "report")[:31]

    # Top-of-sheet metadata block (above the headered table).
    ws.append([f"Domain: {domain['name']}"])
    ws.append([f"Domain ID: {domain['id']}"])
    state_desc = ", ".join(vm_states) if vm_states else "(all states)"
    ws.append([f"State filter: {state_desc}"])
    if days is not None:
        ws.append([f"Days filter: no lifecycle event in last {days} day(s)"])
    ws.append([f"Lifecycle actions: {', '.join(core.LIFECYCLE_ACTIONS)}"])
    ws.append([f"Generated: {datetime.utcnow().isoformat(timespec='seconds')}Z"])
    ws.append([])

    headers = [
        "project_name", "project_id", "instance_uuid", "instance_name",
        "compute_host", "vm_state", "power_state",
        "last_action", "last_action_time", "last_action_user",
        "age_days", "age", "created_at",
    ]
    header_row = ws.max_row + 1
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="left")

    proj_name_by_id = {p["id"]: p["name"] for p in projects}

    def _sort_key(r: Dict[str, Any]):
        return (
            proj_name_by_id.get(r.get("project_id"), "") or "",
            -(r.get("age_seconds") or 0),  # oldest first within a project
        )

    for r in sorted(rows, key=_sort_key):
        age_days = None
        if r.get("age_seconds") is not None:
            age_days = round(r["age_seconds"] / 86400, 2)
        ws.append([
            r.get("project_name") or proj_name_by_id.get(r.get("project_id"), ""),
            r.get("project_id"),
            r.get("uuid"),
            r.get("name"),
            r.get("compute_host"),
            r.get("vm_state"),
            r.get("power_state"),
            r.get("last_action"),
            r.get("last_action_time"),
            r.get("last_action_user"),
            age_days,
            r.get("age"),
            r.get("created_at"),
        ])

    last_row = ws.max_row
    last_col_letter = get_column_letter(len(headers))

    if last_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Approximate auto-fit column widths.
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        values = [ws.cell(row=r, column=col_idx).value for r in range(header_row, last_row + 1)]
        max_len = max((len(str(v)) for v in values if v is not None), default=10)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


if __name__ == "__main__":
    host = os.environ.get("QLR_HOST", "127.0.0.1")
    port = int(os.environ.get("QLR_PORT", "8000"))
    app.run(host=host, port=port, debug=False)
